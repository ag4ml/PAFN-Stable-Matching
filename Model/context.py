import openpyxl
import os
import collections
from Model.advisee import Advisee
from Model.advisor import Advisor
from Model.advisee_mapping import excel_mapping as advisee_excel
from Model.advisor_mapping import excel_mapping as advisor_excel
from Model.advisee_mapping import excel_display_ordered as advisee_display

class Context:
    def __init__(self):
        self.advisors = []
        self.advisees = []
        self.advisor_file = None
        self.advisee_file = None
        self.advisee_file2 = None
        self.setup_filepath()
        self.populate_advisors()
        self.populate_advisees()
        self.insert_advisee_display()

    def setup_filepath(self):
        cur_path = os.path.dirname(__file__)
        self.advisor_file = os.path.join(cur_path, "..", "advisor_bios.xlsx")
        self.advisee_file = os.path.join(cur_path, "..", "advisee_bios.xlsx")
        self.advisee_file2 = os.path.join(cur_path, "..", "advisee_bios_full.xlsx")

    def populate_advisors(self):
        wb = openpyxl.load_workbook(self.advisor_file)
        ws = wb.get_sheet_by_name('Form Responses 1')
        for row in range(2, ws.max_row+1):
            properties = {}
            fname = None
            lname = None
            email = None
            id = row
            for property_name, prop in advisor_excel.items():
                property_column = prop[0]
                property_value = ws[property_column+str(row)].value
                property_value = property_value if property_value is not None else ""
                property_value = property_value.strip() if type(property_value)==str else  property_value
                if property_name=="Email":
                    email = property_value
                elif property_name=="First Name":
                    fname = property_value.title()
                elif property_name=="Last Name":
                    lname = property_value.title()
                else:
                    properties[property_name] = (property_value, prop[1])

            name = fname if fname != "Stephanie" else (fname + " "+ lname[0])
            quota = 5
            new_advisor = Advisor(properties, name,email, id, quota)
            self.advisors.append(new_advisor)

    def populate_advisees(self):
        wb = openpyxl.load_workbook(self.advisee_file)
        ws = wb.get_sheet_by_name('Form Responses 1')
        for row in range(2, ws.max_row+1):
            properties = {}
            name = None
            email = None
            id = row
            ranking = [None, None, None, None, None]
            for property_name, prop in advisee_excel.items():
                property_column = prop[0]
                property_value = ws[property_column+str(row)].value
                property_value = property_value if property_value is not None else ""
                property_value = property_value.strip() if type(property_value)==str else  property_value
                if property_name=="Email":
                    email = property_value
                elif property_name=="Name":
                    name = property_value.title()
                elif property_name in ["Rank1", "Rank2", "Rank3", "Rank4", "Rank5"]:
                    advisor_obj = self.get_advisor(property_value)
                    if(advisor_obj==None): raise Exception("Required advisor not found")
                    ranking[int(property_name[-1])-1]= advisor_obj
                else:
                    properties[property_name]= (property_value, prop[1])
            new_advisee = Advisee(properties, name, email, id, ranking)
            missing_advisors = self.get_missing_advisors(new_advisee.preferred_ranking)
            new_advisee.build_secondary(missing_advisors)
            self.advisees.append(new_advisee)

    def insert_advisee_display(self):
        wb = openpyxl.load_workbook(self.advisee_file2)
        ws = wb.get_sheet_by_name('Form Responses 1')
        for row in range(2, ws.max_row+1):
            properties = collections.OrderedDict()
            for prop_name, prop in advisee_display.items():
                prop_column = prop[0]
                property_value = ws[prop_column+str(row)].value
                property_value = property_value if property_value is not None else ""
                property_value = property_value.strip() if type(property_value)==str else  property_value
                properties[prop_name] = (property_value, prop[1])
            advisee_obj = self.get_advisee(properties["Email"][0])
            advisee_obj.set_display_properties(properties)

    def get_advisor(self, adv_name):
        for advisor in self.advisors:
            if advisor.name==adv_name: return advisor

    def get_advisee(self, adv_email):
        for advisee in self.advisees:
            if advisee.email==adv_email: return advisee

    def get_missing_advisors(self, preferred_ranking):
        adv_names = [advisor.name for advisor in preferred_ranking]
        return [advisor for advisor in self.advisors if advisor.name not in adv_names]

    def has_unmatched_advisees(self):
        for advisee in self.advisees:
            if not advisee.matched[0]: return True
        return False

    def get_unmatched_advisees(self):
        return [advisee for advisee in self.advisees if not advisee.matched[0]]

    def unmatch_adv_pairings(self):
        for advisor in self.advisors:
            advisor.unmatch_unqualified()

    def print_pairings(self):
        for advisor in self.advisors:
            print(advisor)
        advisees_got_top5 = 0
        advisees_got_top2 = 0
        non_ideal = []
        for advisee in self.advisees:
            matched_adv_name = advisee.matched[1]
            top_five = advisee.preferred_ranking
            top_five_names = [advisor.name for advisor in top_five]
            print("{}'s top five are {}".format(advisee.name, top_five_names))
            if matched_adv_name in top_five_names:
                advisees_got_top5+=1
                print("\t He/she got their pick")
            else:
                non_ideal.append(advisee)
            if matched_adv_name in top_five_names[0:2]:
                advisees_got_top2+=1
        print("Percent of advisees with top 5 pick: {}%".format(100*(advisees_got_top5/len(self.advisees))))
        print("Percent of advisees with top 2 pick: {}%".format(100*(advisees_got_top2/len(self.advisees))))
        print("---- Non ideal pairings -----")
        for adv in non_ideal:
            print(adv)
        self.save_to_file()

    def save_to_file(self):
        for advisee in self.advisees:
            if advisee.display_properties=={}: raise NotImplementedError("Doesn't have display: {}".format(advisee.name))
        wb = openpyxl.Workbook()
        dest_filename = "matchings.xlsx"
        ws = wb.active
        ws.title = "Advisor-Advisee pairings"
        n2a = openpyxl.cell.get_column_letter
        row = 1
        for advisor in self.advisors:
            col = 1
            ws[n2a(col)+str(row)] = advisor.name
            ws[n2a(col+1)+str(row)] = advisor.email
            ws[n2a(col+2)+str(row)] = advisor.properties["Major(s)"][0]
            row+=1
            for advisee, advisee_rating in advisor.pairings:
                col=1
                ranking_names = [advisor.name for advisor in advisee.preferred_ranking]
                ws[n2a(col)+str(row)] = (ranking_names.index(advisor.name)+1) if advisor.name in ranking_names else -1
                col = 2
                for prop_name, prop in advisee.display_properties.items():
                    ws[n2a(col)+str(row)] = "{}: {}".format(prop_name, "N/A" if len(prop[0])<3 else prop[0])
                    col+=1
                row+=1
            row+=1
        wb.save(filename=dest_filename)
